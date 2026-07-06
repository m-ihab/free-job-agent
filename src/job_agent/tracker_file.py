"""Import status edits from the internship tracking workbook back into the DB.

Older exporter versions were one-way: every export cleared the data rows and
rewrote them from the database, so a status the user typed into Excel could be
lost on the next export. The exporter now preserves manual columns, and
``import_tracker`` closes that
loop — it reads the workbook, matches each row back to a tracked job (by link,
then by company+title), and applies any changed status to the database via the
tracker. That makes the workbook a genuine two-way surface: edit in Excel, import
to sync, export to refresh.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from job_agent.config import AppConfig
from job_agent.db.database import Database
from job_agent.exporters.internship_workbook import EXPORT_COLUMNS, _default_workbook_path
from job_agent.exporters.internship_workbook_values import normalise_text
from job_agent.exporters.workbook_preserve import sheet_and_header_row
from job_agent.schemas.job import JobListing, JobStatus
from job_agent.tracker import ApplicationTracker

# Human labels (as written by the exporter or typed by the user) → JobStatus.
_STATUS_LABELS: dict[str, JobStatus] = {
    "new": JobStatus.NEW,
    "scored": JobStatus.SCORED,
    "needs review": JobStatus.NEEDS_REVIEW,
    "packet ready": JobStatus.PACKET_READY,
    "applying": JobStatus.APPLYING,
    "needs manual": JobStatus.NEEDS_MANUAL,
    "applied": JobStatus.APPLIED,
    "submitted": JobStatus.MANUALLY_SUBMITTED,
    "manually submitted": JobStatus.MANUALLY_SUBMITTED,
    "auto submitted": JobStatus.AUTO_SUBMITTED,
    "rejected": JobStatus.REJECTED,
    "interview": JobStatus.INTERVIEW,
    "interviewing": JobStatus.INTERVIEW,
    "offer": JobStatus.OFFERED,
    "offered": JobStatus.OFFERED,
    "accepted": JobStatus.ACCEPTED,
    "withdrawn": JobStatus.WITHDRAWN,
    "failed": JobStatus.FAILED,
}


def parse_status_label(label: object) -> Optional[JobStatus]:
    """Map a human/Excel status label to a :class:`JobStatus`, or ``None``."""
    key = normalise_text(label)
    if not key:
        return None
    if key in _STATUS_LABELS:
        return _STATUS_LABELS[key]
    # Accept raw enum values like "MANUALLY_SUBMITTED" or "interview".
    for status in JobStatus:
        if key == status.value.casefold() or key == status.value.replace("_", " ").casefold():
            return status
    return None


def _job_index(jobs: list[JobListing]) -> tuple[dict[str, JobListing], dict[tuple[str, str], JobListing]]:
    by_link: dict[str, JobListing] = {}
    by_company_title: dict[tuple[str, str], JobListing] = {}
    for job in jobs:
        for link in (job.apply_url, job.source_url):
            if link:
                by_link.setdefault(link.strip().casefold(), job)
        by_company_title.setdefault((normalise_text(job.company), normalise_text(job.title)), job)
    return by_link, by_company_title


def import_tracker(config: AppConfig, *, workbook_path: Path | str | None = None) -> dict[str, Any]:
    """Sync status edits from the tracking workbook into the database.

    Returns ``{"updated", "unmatched", "errors", "workbook"}``.
    """
    config.ensure_dirs()
    workbook_file = _default_workbook_path(config, workbook_path)
    if not workbook_file.exists():
        return {"updated": 0, "unmatched": 0, "errors": [f"Workbook not found: {workbook_file}"],
                "workbook": str(workbook_file)}

    tracker = ApplicationTracker(Database(config.db_path))  # type: ignore[arg-type]
    by_link, by_company_title = _job_index(tracker.list_jobs(limit=None))

    workbook = load_workbook(workbook_file, read_only=True, data_only=True)
    worksheet = workbook.active
    header_row, header_map = sheet_and_header_row(worksheet, EXPORT_COLUMNS)
    col = {name: header_map.get(normalise_text(label)) for name, label in (
        ("link", "link to job"), ("company", "company name"),
        ("title", "job title"), ("status", "status"),
    )}

    updated = 0
    unmatched = 0
    errors: list[str] = []

    def _cell(row: tuple, column: int | None) -> str:
        if not column:
            return ""
        idx = column - 1
        if idx < 0 or idx >= len(row):
            return ""
        return str(row[idx].value or "").strip()

    for row in worksheet.iter_rows(min_row=header_row + 1):
        company = _cell(row, col["company"])
        title = _cell(row, col["title"])
        link = _cell(row, col["link"])
        status_label = _cell(row, col["status"])
        if not (company or title or link):
            continue
        status = parse_status_label(status_label)
        if status is None:
            continue
        job = by_link.get(link.casefold()) if link else None
        if job is None:
            job = by_company_title.get((normalise_text(company), normalise_text(title)))
        if job is None:
            unmatched += 1
            continue
        if job.status == status:
            continue
        try:
            tracker.update_status(job.id, status, note="Imported from tracker workbook")
            updated += 1
        except Exception as exc:
            errors.append(f"{company}/{title}: {type(exc).__name__}: {exc}")

    workbook.close()
    return {"updated": updated, "unmatched": unmatched, "errors": errors,
            "workbook": str(workbook_file)}
