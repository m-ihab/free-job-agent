"""Export applied internship applications to an Excel workbook."""
from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from job_agent.config import AppConfig
from job_agent.db.database import Database
from job_agent.exporters.internship_workbook_values import (
    APPLIED_STATUSES,
    applied_at,
    contact_details,
    normalise_text,
    status_label,
)
from job_agent.exporters.workbook_preserve import (
    ensure_headers,
    manual_row_key,
    restore_manual_columns,
    sheet_and_header_row,
    snapshot_manual_columns,
)
from job_agent.intake.internships import is_internship_listing
from job_agent.tracker import ApplicationTracker


# Generic, repo-safe default name. The actual filename used is configurable
# via JOB_AGENT_INTERNSHIP_WORKBOOK so users can keep their private tracker
# under any name they like. Legacy custom files in profiles/ are still
# detected automatically.
DEFAULT_WORKBOOK_NAME = "internship_tracker.xlsx"
LEGACY_WORKBOOK_CANDIDATES = (
    # Names some users had from earlier versions of the project. We still
    # write to them if they exist so no historical data gets orphaned.
    "Internship Search Tracking File A24.xlsx",
    "Internship Tracking.xlsx",
)
EXPORT_COLUMNS = [
    "company name",
    "job title",
    "link to job",
    "job description",
    "location",
    "status",
    "company contact details",
    "date applied",
]
def _default_workbook_path(config: AppConfig, workbook_path: Path | str | None) -> Path:
    """Resolve the destination workbook path.

    Precedence:
    1. Explicit ``workbook_path`` from the caller.
    2. ``JOB_AGENT_INTERNSHIP_WORKBOOK`` env var.
    3. A legacy file in ``profiles/`` if one already exists (don't orphan
       prior user data).
    4. Generic default ``internship_tracker.xlsx``.
    """
    if workbook_path is not None:
        return Path(workbook_path).expanduser()
    env_override = (os.environ.get("JOB_AGENT_INTERNSHIP_WORKBOOK") or "").strip()
    if env_override:
        return Path(env_override).expanduser()
    assert config.profiles_dir is not None
    profiles_dir = Path(config.profiles_dir)
    for legacy in LEGACY_WORKBOOK_CANDIDATES:
        candidate = profiles_dir / legacy
        if candidate.exists():
            return candidate
    return profiles_dir / DEFAULT_WORKBOOK_NAME


def export_applied_internships(
    config: AppConfig,
    *,
    workbook_path: Path | str | None = None,
    sheet_name: str | None = None,
) -> tuple[Path, int]:
    """Write submitted internship applications into the tracking workbook."""
    config.ensure_dirs()
    # ensure_dirs() guarantees db_path is set
    tracker = ApplicationTracker(Database(config.db_path))  # type: ignore[arg-type]
    workbook_file = _default_workbook_path(config, workbook_path)
    workbook_file.parent.mkdir(parents=True, exist_ok=True)

    if workbook_file.exists():
        workbook = load_workbook(workbook_file)
    else:
        workbook = Workbook()

    worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
    header_row, header_map = sheet_and_header_row(worksheet, EXPORT_COLUMNS)
    header_map = ensure_headers(
        worksheet,
        header_row=header_row,
        header_map=header_map,
        exported_columns=EXPORT_COLUMNS,
    )
    manual_snapshot = snapshot_manual_columns(
        worksheet,
        header_row=header_row,
        header_map=header_map,
        exported_columns=EXPORT_COLUMNS,
    )

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        for column_index in range(1, max(worksheet.max_column, *header_map.values()) + 1):
            worksheet.cell(row=row_index, column=column_index).value = None

    jobs = [
        job
        for job in tracker.list_jobs(limit=None)
        if job.status in APPLIED_STATUSES and is_internship_listing(job)
    ]
    jobs.sort(key=lambda job: (job.updated_at, job.created_at), reverse=True)

    for row_offset, job in enumerate(jobs, start=1):
        row_index = header_row + row_offset
        values = {
            "company name": job.company,
            "job title": job.title,
            "link to job": job.apply_url or job.source_url or "",
            "job description": job.description or job.raw_text or "",
            "location": job.location or "",
            "status": status_label(job),
            "company contact details": contact_details(job),
            "date applied": applied_at(tracker, job),
        }
        for column_name, value in values.items():
            worksheet.cell(row=row_index, column=header_map[normalise_text(column_name)], value=value)
        restore_manual_columns(
            worksheet,
            header_row=header_row,
            row=row_index,
            key=manual_row_key(
                job.company,
                job.title,
                job.apply_url or job.source_url or "",
            ),
            snapshot=manual_snapshot,
        )

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for column_name in EXPORT_COLUMNS:
        cell = worksheet.cell(row=header_row, column=header_map[normalise_text(column_name)])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)
    worksheet.auto_filter.ref = f"A{header_row}:{worksheet.cell(row=header_row, column=max(header_map.values())).coordinate}"

    for column_index in range(1, max(header_map.values()) + 1):
        column_letter = worksheet.cell(row=header_row, column=column_index).column_letter
        column_values = [worksheet.cell(row=row, column=column_index).value for row in range(1, worksheet.max_row + 1)]
        max_length = max((len(str(value)) for value in column_values if value is not None), default=0)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    workbook.save(workbook_file)
    return workbook_file, len(jobs)
