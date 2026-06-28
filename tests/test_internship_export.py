from openpyxl import Workbook, load_workbook

from job_agent.db.database import Database
from job_agent.config import AppConfig
from job_agent.exporters.internship_workbook import DEFAULT_WORKBOOK_NAME, EXPORT_COLUMNS, export_applied_internships
from job_agent.schemas.job import JobListing, JobStatus
from job_agent.tracker import ApplicationTracker


def test_export_applied_internships_writes_only_internships(tmp_path):
    profiles_dir = tmp_path / "profiles"
    profiles_dir.mkdir()
    data_dir = tmp_path / "data"
    output_dir = tmp_path / "outputs"
    config = AppConfig(data_dir=data_dir, profiles_dir=profiles_dir, outputs_dir=output_dir)
    Database(config.db_path).initialize()
    tracker = ApplicationTracker(Database(config.db_path))

    internship = JobListing(
        title="Data Science Intern",
        company="ACME",
        description="Join our internship program in Paris.",
        location="Paris",
        apply_url="https://example.com/jobs/1",
        job_type="Internship",
    )
    tracker.add_job(internship)
    tracker.update_status(internship.id, JobStatus.APPLIED, note="submitted")

    regular_job = JobListing(
        title="Data Scientist",
        company="ACME",
        description="Full-time role.",
        location="Paris",
        apply_url="https://example.com/jobs/2",
        job_type="Permanent",
    )
    tracker.add_job(regular_job)
    tracker.update_status(regular_job.id, JobStatus.APPLIED, note="submitted")

    workbook_path, count = export_applied_internships(config)

    assert workbook_path == profiles_dir / DEFAULT_WORKBOOK_NAME
    assert count == 1
    workbook = load_workbook(workbook_path)
    worksheet = workbook.active

    headers = [worksheet.cell(row=1, column=index).value for index in range(1, 9)]
    assert headers == [
        "Company Name",
        "Job Title",
        "Link To Job",
        "Job Description",
        "Location",
        "Status",
        "Company Contact Details",
        "Date Applied",
    ]

    assert worksheet.cell(row=2, column=1).value == "ACME"
    assert worksheet.cell(row=2, column=2).value == "Data Science Intern"
    assert worksheet.cell(row=2, column=6).value == "Applied"
    assert worksheet.cell(row=2, column=8).value
    assert worksheet.cell(row=3, column=1).value is None


def test_export_applied_internships_preserves_manual_workbook_columns(tmp_path):
    profiles_dir = tmp_path / "profiles"
    profiles_dir.mkdir()
    config = AppConfig(data_dir=tmp_path / "data", profiles_dir=profiles_dir, outputs_dir=tmp_path / "outputs")
    Database(config.db_path).initialize()
    tracker = ApplicationTracker(Database(config.db_path))

    job = JobListing(
        title="Data Science Intern",
        company="ACME",
        description="Stage data science in Paris.",
        location="Paris",
        apply_url="https://example.com/jobs/1",
        job_type="Internship",
    )
    tracker.add_job(job)
    tracker.update_status(job.id, JobStatus.APPLIED, note="submitted")

    workbook_path = profiles_dir / DEFAULT_WORKBOOK_NAME
    workbook = Workbook()
    worksheet = workbook.active
    for column, name in enumerate(EXPORT_COLUMNS, start=1):
        worksheet.cell(row=1, column=column, value=name.title())
    worksheet.cell(row=1, column=len(EXPORT_COLUMNS) + 1, value="Personal Notes")
    worksheet.cell(row=1, column=len(EXPORT_COLUMNS) + 2, value="Referral Owner")
    worksheet.cell(row=2, column=1, value="ACME")
    worksheet.cell(row=2, column=2, value="Data Science Intern")
    worksheet.cell(row=2, column=3, value="https://example.com/jobs/1")
    worksheet.cell(row=2, column=len(EXPORT_COLUMNS) + 1, value="Ask Sam for referral")
    worksheet.cell(row=2, column=len(EXPORT_COLUMNS) + 2, value="Sam")
    workbook.save(workbook_path)

    export_applied_internships(config)

    reloaded = load_workbook(workbook_path)
    sheet = reloaded.active
    assert sheet.cell(row=1, column=9).value == "Personal Notes"
    assert sheet.cell(row=1, column=10).value == "Referral Owner"
    assert sheet.cell(row=2, column=1).value == "ACME"
    assert sheet.cell(row=2, column=9).value == "Ask Sam for referral"
    assert sheet.cell(row=2, column=10).value == "Sam"
