"""TDD: auto-export the internship workbook whenever a job is marked submitted.

RED: these tests currently fail because mark-submitted does not trigger
     an automatic export.  GREEN: add the auto-export call to both the
     CLI handler and the UI /api/status endpoint.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from click.testing import CliRunner

from job_agent.cli.main import app
from job_agent.config import AppConfig
from job_agent.db.database import Database
from job_agent.exporters.internship_workbook import DEFAULT_WORKBOOK_NAME
from job_agent.schemas.job import JobListing, JobStatus
from job_agent.schemas.packet import ApplicationPacket, PacketStatus
from job_agent.tracker import ApplicationTracker


# ── helpers ──────────────────────────────────────────────────────────────────

def _make_internship(tracker: ApplicationTracker) -> JobListing:
    job = JobListing(
        title="Data Science Intern",
        company="Acme",
        description="Stage data science à Paris.",
        location="Paris",
        apply_url="https://example.com/apply/1",
        job_type="Internship",
    )
    tracker.add_job(job)
    return job


def _make_packet(db: Database, job_id: str) -> ApplicationPacket:
    packet = ApplicationPacket(
        job_id=job_id,
        status=PacketStatus.READY,
        fit_score=80.0,
        fit_decision="apply",
        cover_letter_md="Dear Hiring Manager,\n\nI am a great fit.",
    )
    db.save_packet(packet)
    return packet


def _env(tmp_path: Path) -> dict[str, str]:
    return {
        "JOB_AGENT_DATA_DIR": str(tmp_path / "data"),
        "JOB_AGENT_PROFILES_DIR": str(tmp_path / "profiles"),
        "JOB_AGENT_OUTPUTS_DIR": str(tmp_path / "outputs"),
    }


# ── CLI: mark-submitted → workbook auto-created ───────────────────────────────

class TestCliMarkSubmittedAutoExport:
    def test_workbook_created_after_mark_submitted(self, monkeypatch, tmp_path: Path) -> None:
        monkeypatch.setenv("JOB_AGENT_DATA_DIR", str(tmp_path / "data"))
        monkeypatch.setenv("JOB_AGENT_PROFILES_DIR", str(tmp_path / "profiles"))
        monkeypatch.setenv("JOB_AGENT_OUTPUTS_DIR", str(tmp_path / "outputs"))
        config = AppConfig(
            data_dir=tmp_path / "data",
            profiles_dir=tmp_path / "profiles",
            outputs_dir=tmp_path / "outputs",
        )
        config.ensure_dirs()
        db = Database(config.db_path)
        db.initialize()
        tracker = ApplicationTracker(db)
        job = _make_internship(tracker)
        packet = _make_packet(db, job.id)

        runner = CliRunner()
        result = runner.invoke(app, ["mark-submitted", packet.id])

        assert result.exit_code == 0, result.output
        workbook = tmp_path / "profiles" / DEFAULT_WORKBOOK_NAME
        assert workbook.exists(), "Internship tracker workbook should be created automatically"

    def test_workbook_contains_submitted_internship(self, monkeypatch, tmp_path: Path) -> None:
        from openpyxl import load_workbook as lw

        monkeypatch.setenv("JOB_AGENT_DATA_DIR", str(tmp_path / "data"))
        monkeypatch.setenv("JOB_AGENT_PROFILES_DIR", str(tmp_path / "profiles"))
        monkeypatch.setenv("JOB_AGENT_OUTPUTS_DIR", str(tmp_path / "outputs"))
        config = AppConfig(
            data_dir=tmp_path / "data",
            profiles_dir=tmp_path / "profiles",
            outputs_dir=tmp_path / "outputs",
        )
        config.ensure_dirs()
        db = Database(config.db_path)
        db.initialize()
        tracker = ApplicationTracker(db)
        job = _make_internship(tracker)
        packet = _make_packet(db, job.id)

        runner = CliRunner()
        runner.invoke(app, ["mark-submitted", packet.id])

        workbook = tmp_path / "profiles" / DEFAULT_WORKBOOK_NAME
        wb = lw(workbook)
        ws = wb.active
        cell_values = [ws.cell(row=row, column=1).value for row in range(1, ws.max_row + 1)]
        assert "Acme" in cell_values, "Company name should appear in the tracker workbook"

    def test_workbook_not_created_for_non_internship(self, monkeypatch, tmp_path: Path) -> None:
        monkeypatch.setenv("JOB_AGENT_DATA_DIR", str(tmp_path / "data"))
        monkeypatch.setenv("JOB_AGENT_PROFILES_DIR", str(tmp_path / "profiles"))
        monkeypatch.setenv("JOB_AGENT_OUTPUTS_DIR", str(tmp_path / "outputs"))
        config = AppConfig(
            data_dir=tmp_path / "data",
            profiles_dir=tmp_path / "profiles",
            outputs_dir=tmp_path / "outputs",
        )
        config.ensure_dirs()
        db = Database(config.db_path)
        db.initialize()
        tracker = ApplicationTracker(db)
        job = JobListing(
            title="Senior Data Scientist",
            company="BigCo",
            description="Full-time permanent role.",
            apply_url="https://example.com/apply/2",
            job_type="Permanent",
        )
        tracker.add_job(job)
        packet = _make_packet(db, job.id)

        runner = CliRunner()
        runner.invoke(app, ["mark-submitted", packet.id])

        workbook = tmp_path / "profiles" / DEFAULT_WORKBOOK_NAME
        if workbook.exists():
            from openpyxl import load_workbook as lw
            wb = lw(workbook)
            ws = wb.active
            data_rows = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
            assert "BigCo" not in data_rows, "Non-internship job should not appear in tracker"
