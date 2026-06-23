"""Tests for importing status edits from the tracking workbook into the DB."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from job_agent.config import AppConfig
from job_agent.db.database import Database
from job_agent.exporters.internship_workbook import EXPORT_COLUMNS, _default_workbook_path
from job_agent.schemas.job import JobListing, JobStatus
from job_agent.tracker import ApplicationTracker
from job_agent.tracker_file import import_tracker, parse_status_label


def _config(tmp_path: Path) -> AppConfig:
    profiles_dir = tmp_path / "profiles"
    profiles_dir.mkdir()
    config = AppConfig(data_dir=tmp_path / "data", profiles_dir=profiles_dir,
                       outputs_dir=tmp_path / "outputs")
    Database(config.db_path).initialize()
    return config


def _write_workbook(config: AppConfig, rows: list[dict]) -> Path:
    path = _default_workbook_path(config, None)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    for col, name in enumerate(EXPORT_COLUMNS, start=1):
        worksheet.cell(row=1, column=col, value=name.title())
    index = {name: col for col, name in enumerate(EXPORT_COLUMNS, start=1)}
    for offset, row in enumerate(rows, start=2):
        for key, value in row.items():
            worksheet.cell(row=offset, column=index[key], value=value)
    workbook.save(path)
    return path


class TestParseStatusLabel:
    def test_common_labels(self) -> None:
        assert parse_status_label("Interview") is JobStatus.INTERVIEW
        assert parse_status_label("Applied") is JobStatus.APPLIED
        assert parse_status_label("rejected") is JobStatus.REJECTED

    def test_raw_enum_value(self) -> None:
        assert parse_status_label("MANUALLY_SUBMITTED") is JobStatus.MANUALLY_SUBMITTED

    def test_unknown_returns_none(self) -> None:
        assert parse_status_label("banana") is None
        assert parse_status_label("") is None


class TestImportTracker:
    def test_status_edit_syncs_to_db(self, tmp_path) -> None:
        config = _config(tmp_path)
        tracker = ApplicationTracker(Database(config.db_path))
        job = JobListing(title="Data Science Intern", company="ACME",
                         apply_url="https://example.com/jobs/1", location="Paris")
        tracker.add_job(job)

        _write_workbook(config, [{
            "company name": "ACME", "job title": "Data Science Intern",
            "link to job": "https://example.com/jobs/1", "status": "Interview",
        }])
        result = import_tracker(config)

        assert result["updated"] == 1
        assert result["unmatched"] == 0
        assert tracker.get_job(job.id).status is JobStatus.INTERVIEW

    def test_matches_by_company_and_title_without_link(self, tmp_path) -> None:
        config = _config(tmp_path)
        tracker = ApplicationTracker(Database(config.db_path))
        job = JobListing(title="ML Engineer", company="Beta", location="Lyon")
        tracker.add_job(job)

        _write_workbook(config, [{
            "company name": "beta", "job title": "ml engineer", "status": "Rejected",
        }])
        result = import_tracker(config)

        assert result["updated"] == 1
        assert tracker.get_job(job.id).status is JobStatus.REJECTED

    def test_unmatched_row_is_counted(self, tmp_path) -> None:
        config = _config(tmp_path)
        _write_workbook(config, [{
            "company name": "Nobody", "job title": "Ghost", "status": "Applied",
        }])
        result = import_tracker(config)
        assert result["updated"] == 0
        assert result["unmatched"] == 1

    def test_unknown_status_is_skipped(self, tmp_path) -> None:
        config = _config(tmp_path)
        tracker = ApplicationTracker(Database(config.db_path))
        job = JobListing(title="Data Analyst", company="Gamma")
        tracker.add_job(job)
        _write_workbook(config, [{
            "company name": "Gamma", "job title": "Data Analyst", "status": "??unknown??",
        }])
        result = import_tracker(config)
        assert result["updated"] == 0
        assert tracker.get_job(job.id).status is JobStatus.NEW

    def test_missing_workbook_reports_error(self, tmp_path) -> None:
        config = _config(tmp_path)
        result = import_tracker(config)
        assert result["updated"] == 0
        assert result["errors"]
